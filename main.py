import time

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By

# Abrir o site da empresa para a verificação dos pagamentos via cpf
driver = webdriver.Chrome()
driver.get("https://consultcpf-devaprender.netlify.app")
time.sleep(3.5)

# Abrir a minha planilha
planilha_clientes = openpyxl.load_workbook('dados_clientes.xlsx')
pagina_clientes = planilha_clientes['Sheet1']

# loop para ler todos os dados dos clientes
for linhas in pagina_clientes.iter_rows(min_row=2, values_only=True):
    # Dados armazenados em seus respectivos lugares
    nome, valor, cpf, data = linhas

    # Verificar o cpf
    verificador = driver.find_element(By.XPATH, "//input[@id='cpfInput']")
    verificador.clear()
    verificador.send_keys(cpf)

    time.sleep(0.5)

    # Clickar no botão
    consultar = driver.find_element(By.XPATH, "//button[@class='btn btn-custom btn-lg btn-block mt-3']").click()
    time.sleep(3.5)
    status = driver.find_element(By.XPATH, "//span[@id='statusLabel']")
    if status.text == 'em dia':
        # Abrir a minha planilha de fechamento para armazenar os status dos clientes
        planilha_fechamento = openpyxl.load_workbook('planilha fechamento.xlsx')
        pagina_fechamento = planilha_fechamento['Sheet1']

        # Pegar a data de pagamento de cada cliente
        texto_data = driver.find_element(By.XPATH, "//p[@id='paymentDate']").text.split()
        data_pagamento = texto_data[3]

        # Pegar o metodo de pagamento dos clientes
        texto_metodo = driver.find_element(By.XPATH, "//p[@id='paymentMethod']").text.split()
        metodo = texto_metodo[3]
        pagina_fechamento.append([nome, valor, cpf, data, 'em dia', data_pagamento,metodo])

        # Salvar meus novos dados na minha planilha de fechamento
        planilha_fechamento.save('planilha fechamento.xlsx')


    else:
        # Abrir a minha planilha
        planilha_fechamento = openpyxl.load_workbook('planilha fechamento.xlsx')
        pagina_fechamento = planilha_fechamento['Sheet1']
        pagina_fechamento.append([nome, valor, cpf,data,'Pendente'])


        planilha_fechamento.save('planilha fechamento.xlsx')

